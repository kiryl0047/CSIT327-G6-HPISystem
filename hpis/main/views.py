from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.utils.crypto import get_random_string
import string
from django.contrib import messages
from django.http import JsonResponse, FileResponse
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from datetime import timedelta
import json
import csv
from django.contrib.auth import update_session_auth_hash
from io import BytesIO, StringIO
from functools import wraps

from .forms import LoginForm, CustomPasswordChangeForm, NotificationPreferencesForm, UserProfileForm
from .models import (
    UserProfile, NotificationPreference, AccessLog,
    DataExportRequest, DeleteAccountRequest, PatientAppointment, Report
)
from django.db import transaction
from django.core.mail import send_mail
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

from django.contrib.auth import logout
from django.shortcuts import redirect
from django.views.decorators.http import require_POST

from django.views.decorators.cache import never_cache


@require_POST
def logout_view(request):
    logout(request)
    request.session.flush()
    return redirect('user_login')


# ==================== Decorators ====================

@never_cache
@login_required
def superadmin_dashboard(request):
    ...

@never_cache
@login_required
def admin_dashboard(request):
    ...

def get_client_ip(request):
    """Extract client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def log_access(user, access_type, description="", request=None):
    """Log user access for security tracking"""
    ip_address = get_client_ip(request) if request else None
    user_agent = request.META.get('HTTP_USER_AGENT', '') if request else ''

    AccessLog.objects.create(
        user=user,
        access_type=access_type,
        ip_address=ip_address,
        user_agent=user_agent,
        description=description
    )


def role_required(*allowed_roles):
    """Decorator to check if user has required role"""

    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('user_login')

            try:
                user_role = request.user.profile.role
            except UserProfile.DoesNotExist:
                return redirect('user_login')

            if user_role not in allowed_roles:
                messages.error(request, 'You do not have permission to access this page.')
                return redirect('homepage')

            return view_func(request, *args, **kwargs)

        return wrapper

    return decorator


# ==================== Authentication Views ====================

def landing_page(request):
    """Landing page with appointment booking form for patients"""
    if request.method == "POST":
        # Create appointment from form data
        appointment = PatientAppointment.objects.create(
            first_name=request.POST.get('first_name'),
            last_name=request.POST.get('last_name'),
            middle_name=request.POST.get('middle_name', ''),
            date_of_birth=request.POST.get('dob'),
            gender=request.POST.get('gender'),
            email=request.POST.get('email'),
            contact_number=request.POST.get('contact'),
            address=request.POST.get('address'),
            appointment_type=request.POST.get('appointment_type'),
            appointment_date=request.POST.get('available_date'),
            notes=request.POST.get('notes', ''),
            status='pending'
        )

        messages.success(
            request,
            'Your appointment request has been submitted. An admin will review and assign it to a doctor shortly.'
        )
        return render(request, 'landing_page.html', {'appointment_submitted': True})

    return render(request, 'landing_page.html')


def user_login(request):
    """Handle user login - Staff only"""
    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            identifier = form.cleaned_data['username']
            password = form.cleaned_data['password']

            try:
                user_obj = User.objects.get(username__iexact=identifier)
            except User.DoesNotExist:
                messages.error(request, "User not found.")
                return render(request, "login.html", {"form": form})

            user = authenticate(request, username=user_obj.username, password=password)

            if user is not None:
                login(request, user)
                # If the user's profile requires a password change, redirect them to change-password
                try:
                    if user.profile.must_change_password:
                        return redirect('change_password')
                except Exception:
                    pass

                log_access(user, 'login', 'User logged in', request)

                if not request.POST.get("remember_me"):
                    request.session.set_expiry(0)
                else:
                    request.session.set_expiry(1209600)

                # Redirect based on Django flags first (superuser/staff), then profile.role
                try:
                    if user.is_superuser:
                        return redirect('super_admin_dashboard')
                    if user.is_staff:
                        return redirect('admin_dashboard')

                    role = user.profile.role
                    if role == 'super_admin':
                        return redirect('super_admin_dashboard')
                    elif role == 'admin':
                        return redirect('admin_dashboard')
                    elif role == 'doctor':
                        return redirect('doctor_dashboard')
                except UserProfile.DoesNotExist:
                    return redirect('homepage')
            else:
                messages.error(request, "Incorrect password.")
                if 'user_obj' in locals():
                    log_access(user_obj, 'failed_login', 'Failed login attempt', request)
    else:
        form = LoginForm()

    return render(request, "login.html", {"form": form})


# ==================== Super Admin Dashboard ====================

@login_required
@role_required('super_admin')
def super_admin_dashboard(request):
    """Super Admin main dashboard"""
    from django.db.models import Count, Q
    
    # Optimize with single aggregated query
    staff_counts = UserProfile.objects.aggregate(
        total_staff=Count('id', filter=Q(role__in=['admin', 'doctor'])),
        admins=Count('id', filter=Q(role='admin')),
        doctors=Count('id', filter=Q(role='doctor')),
    )
    
    stats = {
        'total_staff': staff_counts['total_staff'],
        'admins': staff_counts['admins'],
        'doctors': staff_counts['doctors'],
        'pending_appointments': PatientAppointment.objects.filter(status='pending').count(),
    }

    log_access(request.user, 'data_view', 'Accessed super admin dashboard', request)

    return render(request, 'super_admin_dashboard.html', {'stats': stats})


@login_required
@role_required('super_admin')
def create_staff(request):
    """Create new staff account (Admin or Doctor)"""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        role = request.POST.get('role')
        department = request.POST.get('department', '')
        license_number = request.POST.get('license_number', '')

        # Validate role
        if role not in ['admin', 'doctor']:
            messages.error(request, 'Invalid role.')
            return redirect('create_staff')

        # Check if user exists
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return redirect('create_staff')

        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists.')
            return redirect('create_staff')

        try:
            # Create user with a secure temporary password (letters+digits+punctuation)
            alphabet = string.ascii_letters + string.digits + string.punctuation
            temp_password = get_random_string(12, alphabet)
            # Create user with temporary password
            # temp_password = User.objects.make_random_password(12)
            temp_password = str(123456)
            user = User.objects.create_user(
                username=username,
                email=email,
                password=temp_password,
                first_name=first_name,
                last_name=last_name
            )

            # Create profile with role and notification preferences atomically
            # Read optional checkbox to email temp password
            email_temp = request.POST.get('email_temp_password') == 'on'

            with transaction.atomic():
                # Safely get the profile (which may have been created by a signal)
                profile, created = UserProfile.objects.get_or_create(user=user)

                # Now, update the profile's fields regardless of whether it was created now or by a signal
                profile.role = role
                profile.department = department
                profile.license_number = license_number

                # Set must_change_password flag
                profile.must_change_password = True

                profile.save() # <-- Save the updated fields

                # Keep the NotificationPreference check as a safety measure
                NotificationPreference.objects.get_or_create(user=user)

            # Optionally send the temporary password via email (HTML)
            if email_temp and email:
                try:
                    subject = 'Welcome to the Healthcare Patient Information System'
                    plain_message = (
                        f'Hello {first_name or username},\n\n'
                        f'An account has been created for you.\n'
                        f'Username: {username}\n'
                        f'Temporary password: {temp_password}\n\n'
                        'Please change your password on first login.\n\n'
                        'If you have any questions, contact IT Support.'
                    )

                    html_message = (
                        f'<p>Hello {first_name or username},</p>'
                        f'<p>An account has been created for you.</p>'
                        f'<ul><li><strong>Username:</strong> {username}</li>'
                        f'<li><strong>Temporary password:</strong> <code>{temp_password}</code></li></ul>'
                        f'<p>Please change your password on first login.</p>'
                        f'<p>If you have any questions, contact IT Support.</p>'
                    )

                    send_mail(
                        subject=subject,
                        message=plain_message,
                        from_email=None,
                        recipient_list=[email],
                        html_message=html_message,
                        fail_silently=True,
                    )
                except Exception:
                    # Swallow email errors but log if desired
                    pass

            log_access(
                request.user,
                'data_update',
                f'Created new {role} account: {username}',
                request
            )

            # Render a success page that shows the temporary password so admin can copy it
            return render(request, 'create_staff_success.html', {
                'username': username,
                'email': email,
                'role': role,
                'temp_password': temp_password,
                'profile': profile,
            })

        except Exception as e:
            messages.error(request, f'Error creating account: {str(e)}')
            return redirect('create_staff')

    return render(request, 'create_staff.html')


@login_required
@role_required('super_admin')
def manage_staff(request):
    """Manage all staff accounts"""
    # Optimize with select_related and only needed fields
    staff = User.objects.filter(
        profile__role__in=['admin', 'doctor']
    ).select_related('profile').only(
        'id', 'username', 'email', 'first_name', 'last_name',
        'profile__role', 'profile__department', 'profile__license_number'
    )

    log_access(request.user, 'data_view', 'Accessed manage staff page', request)

    return render(request, 'manage_staff.html', {'staff': staff})


@login_required
@role_required('super_admin')
def edit_staff(request, staff_id):
    """Edit staff account"""
    staff_user = get_object_or_404(User, id=staff_id)

    if not staff_user.profile or staff_user.profile.role not in ['admin', 'doctor']:
        messages.error(request, 'Invalid staff member.')
        return redirect('manage_staff')

    if request.method == 'POST':
        staff_user.first_name = request.POST.get('first_name')
        staff_user.last_name = request.POST.get('last_name')
        staff_user.email = request.POST.get('email')
        staff_user.save()

        staff_user.profile.role = request.POST.get('role')
        staff_user.profile.department = request.POST.get('department')
        staff_user.profile.license_number = request.POST.get('license_number')
        staff_user.profile.save()

        log_access(
            request.user,
            'data_update',
            f'Updated staff account: {staff_user.username}',
            request
        )

        messages.success(request, 'Staff account updated successfully.')
        return redirect('manage_staff')

    return render(request, 'edit_staff.html', {'staff_user': staff_user})


@login_required
@role_required('super_admin')
def delete_staff(request, staff_id):
    """Delete staff account"""
    staff_user = get_object_or_404(User, id=staff_id)

    if request.method == 'POST':
        username = staff_user.username
        staff_user.delete()

        log_access(
            request.user,
            'data_update',
            f'Deleted staff account: {username}',
            request
        )

        messages.success(request, f'Staff account {username} deleted.')
        return redirect('manage_staff')

    return render(request, 'delete_staff_confirm.html', {'staff_user': staff_user})


# ==================== Admin Dashboard ====================

@login_required
@role_required('admin')
def admin_dashboard(request):
    """Admin main dashboard"""
    # Optimize queries with select_related to avoid N+1 queries
    pending_appointments = PatientAppointment.objects.filter(status='pending').select_related(
        'assigned_doctor', 'assigned_admin'
    )[:20]  # Limit to recent 20
    
    assigned_appointments = PatientAppointment.objects.filter(status='assigned').select_related(
        'assigned_doctor', 'assigned_admin'
    )[:20]  # Limit to recent 20
    
    doctors = User.objects.filter(profile__role='doctor').select_related('profile')

    # Use only() to fetch only needed fields for counts
    pending_count = PatientAppointment.objects.filter(status='pending').count()
    assigned_count = PatientAppointment.objects.filter(status='assigned').count()

    log_access(request.user, 'data_view', 'Accessed admin dashboard', request)

    context = {
        'pending_count': pending_count,
        'assigned_count': assigned_count,
        'pending_appointments': pending_appointments,
        'assigned_appointments': assigned_appointments,
        'doctors': doctors,
    }

    return render(request, 'admin_dashboard.html', context)


@login_required
@role_required('admin')
def assign_appointment(request, appointment_id):
    """Assign appointment to a doctor"""
    appointment = get_object_or_404(PatientAppointment, id=appointment_id)

    if request.method == 'POST':
        doctor_id = request.POST.get('doctor_id')
        appointment_date = request.POST.get('appointment_date')
        appointment_time = request.POST.get('appointment_time')

        try:
            doctor = User.objects.get(id=doctor_id, profile__role='doctor')

            appointment.assigned_doctor = doctor
            appointment.assigned_admin = request.user
            appointment.status = 'assigned'
            appointment.appointment_date = appointment_date
            if appointment_time:
                appointment.appointment_time = appointment_time
            appointment.save()

            log_access(
                request.user,
                'data_update',
                f'Assigned appointment to Dr. {doctor.first_name}',
                request
            )

            messages.success(request, 'Appointment assigned successfully.')
            return redirect('admin_dashboard')
        except User.DoesNotExist:
            messages.error(request, 'Doctor not found.')

    doctors = User.objects.filter(profile__role='doctor').select_related('profile').only(
        'id', 'first_name', 'last_name', 'email'
    )
    return render(request, 'assign_appointment.html', {'appointment': appointment, 'doctors': doctors})


# ==================== Doctor Dashboard ====================

@login_required
@role_required('doctor')
def doctor_dashboard(request):
    """Doctor main dashboard"""
    # Optimize with select_related and limit results
    my_appointments = PatientAppointment.objects.filter(
        assigned_doctor=request.user
    ).select_related('assigned_admin')[:50]  # Limit to recent 50

    # Use separate counts to avoid loading all data
    confirmed_count = PatientAppointment.objects.filter(
        assigned_doctor=request.user, status='confirmed'
    ).count()
    pending_count = PatientAppointment.objects.filter(
        assigned_doctor=request.user, status='assigned'
    ).count()

    log_access(request.user, 'data_view', 'Accessed doctor dashboard', request)

    context = {
        'appointments': my_appointments,
        'confirmed_count': confirmed_count,
        'pending_count': pending_count,
    }

    return render(request, 'doctor_dashboard.html', context)


@login_required
@role_required('doctor')
def confirm_appointment(request, appointment_id):
    """Doctor confirms appointment"""
    appointment = get_object_or_404(PatientAppointment, id=appointment_id)

    if appointment.assigned_doctor != request.user:
        messages.error(request, 'You cannot confirm this appointment.')
        return redirect('doctor_dashboard')

    if request.method == 'POST':
        appointment.status = 'confirmed'
        appointment.save()

        log_access(
            request.user,
            'data_update',
            f'Confirmed appointment for {appointment.first_name}',
            request
        )

        messages.success(request, 'Appointment confirmed.')
        return redirect('doctor_dashboard')

    return render(request, 'confirm_appointment.html', {'appointment': appointment})


# ==================== General User Views ====================

@login_required
def homepage(request):
    """User homepage - redirects based on role"""
    try:
        role = request.user.profile.role
        if role == 'super_admin':
            return redirect('super_admin_dashboard')
        elif role == 'admin':
            return redirect('admin_dashboard')
        elif role == 'doctor':
            return redirect('doctor_dashboard')
    except UserProfile.DoesNotExist:
        pass

    return redirect('user_login')


# ==================== Settings Views ====================

@login_required
def profile(request):
    """User profile page"""
    try:
        user_profile = request.user.profile
    except UserProfile.DoesNotExist:
        user_profile = UserProfile.objects.create(user=request.user)

    context = {
        'user_profile': user_profile,
    }

    return render(request, 'profile.html', context)


@login_required
def settings(request):
    """User settings page"""
    try:
        user_profile = request.user.profile
    except UserProfile.DoesNotExist:
        user_profile = UserProfile.objects.create(user=request.user)

    try:
        notification_pref = request.user.notification_preference
    except NotificationPreference.DoesNotExist:
        notification_pref = NotificationPreference.objects.create(user=request.user)

    # Limit and optimize access logs query
    access_logs = AccessLog.objects.filter(user=request.user).only(
        'access_type', 'ip_address', 'timestamp', 'description'
    ).order_by('-timestamp')[:50]

    context = {
        'user_profile': user_profile,
        'notification_pref': notification_pref,
        'access_logs': access_logs,
    }

    return render(request, 'settings.html', context)


@login_required
# REMOVED: @require_http_methods(["POST"])
@require_http_methods(["GET", "POST"])
def update_profile(request):
    """Handle user profile updates"""
    try:
        user_profile = request.user.profile
    except UserProfile.DoesNotExist:
        user_profile = UserProfile.objects.create(user=request.user)

    if request.method == 'POST':
        request.user.first_name = request.POST.get('first_name')
        request.user.last_name = request.POST.get('last_name')
        request.user.email = request.POST.get('email')
        request.user.save()

        log_access(
            request.user,
            'data_update',
            'Updated profile information',
            request
        )

        messages.success(request, 'Profile updated successfully!')
        # Redirect to profile page if coming from there, otherwise settings
        return redirect(request.GET.get('next', 'profile'))

    return render(request, 'settings.html')


@login_required
@require_http_methods(["GET", "POST"])
def change_password(request):
    """
    Handle password change.
    GET: Displays the form.
    POST: Processes the form submission.
    """
    if request.method == 'POST':
        form = CustomPasswordChangeForm(request.user, request.POST)

        if form.is_valid():
            user = form.save()

            # 🌟 FIX: Keep the user logged in after password change 🌟
            update_session_auth_hash(request, user)
            update_session_auth_hash(request, user)

            # Clear must_change_password flag if present...
            # ... (rest of your existing code for clearing the flag)
            try:
                profile = request.user.profile
                if profile.must_change_password:
                    profile.must_change_password = False
                    profile.save()
            except Exception:
                pass
            # ... (rest of your existing code)

            log_access(
                request.user,
                'data_update',
                'Changed password',
                request
            )

            messages.success(request, 'Password changed successfully! You are still logged in.')
            return redirect('settings')

    else: # Handles the GET request, including the redirect from user_login
        form = CustomPasswordChangeForm(request.user)

    # Renders the change password form, likely integrated into settings.html
    # Note: If your change_password URL is meant to be a standalone page,
    #       you might need a separate template, but based on your structure,
    #       it seems to be a dedicated endpoint that renders the form within settings.
    # We will redirect to settings instead, as the form should live there.

    # If the user is on the change_password URL directly via GET, we redirect them
    # to the main settings page which will render the form using the settings_page view.
    return redirect('settings')

# The settings_page view needs to be updated to pass the form object for the template.
# But for now, fixing the redirect logic in user_login is the most direct solution
# combined with removing the POST-only decorator.


@login_required
@require_http_methods(["POST"])
def update_notifications(request):
    """Handle notification preference updates"""
    try:
        notification_pref = request.user.notification_preference
    except NotificationPreference.DoesNotExist:
        notification_pref = NotificationPreference.objects.create(user=request.user)

    if request.method == 'POST':
        notification_pref.email_notifications = request.POST.get('email_notifications') == 'on'
        notification_pref.sms_notifications = request.POST.get('sms_notifications') == 'on'
        notification_pref.prescription_alerts = request.POST.get('prescription_alerts') == 'on'
        notification_pref.system_updates = request.POST.get('system_updates') == 'on'
        notification_pref.save()

        log_access(
            request.user,
            'data_update',
            'Updated notification preferences',
            request
        )

        messages.success(request, 'Notification preferences updated successfully!')
        return redirect('settings')

    return JsonResponse({'status': 'error'}, status=400)


@login_required
def access_logs(request):
    """Display user access logs"""
    logs = AccessLog.objects.filter(user=request.user).order_by('-timestamp')[:50]

    log_access(
        request.user,
        'data_view',
        'Viewed access logs',
        request
    )

    context = {
        'access_logs': logs
    }

    return render(request, 'settings.html', context)


@login_required
@require_http_methods(["POST"])
def request_data_export(request):
    """Request data export"""

    # Check if there's already a pending request
    existing_request = DataExportRequest.objects.filter(
        user=request.user,
        status__in=['pending', 'processing']
    ).exists()

    if existing_request:
        messages.warning(request, 'You already have a pending data export request.')
        return redirect('settings')

    # Create export request
    export_request = DataExportRequest.objects.create(
        user=request.user,
        status='processing',
        expires_at=timezone.now() + timedelta(days=7)
    )

    log_access(
        request.user,
        'data_download',
        'Requested data export',
        request
    )

    # Generate export immediately
    export_user_data(request.user)

    messages.success(request, 'Data export completed. You can download it now.')
    return redirect('settings')


def export_user_data(user):
    """Generate and store user data export"""
    output = StringIO()
    writer = csv.writer(output)

    writer.writerow(['Personal Information'])
    writer.writerow(['Field', 'Value'])
    writer.writerow(['Username', user.username])
    writer.writerow(['Email', user.email])
    writer.writerow(['First Name', user.first_name])
    writer.writerow(['Last Name', user.last_name])
    writer.writerow(['Joined Date', user.date_joined])

    writer.writerow([])
    writer.writerow(['Account Information'])
    writer.writerow(['Field', 'Value'])

    try:
        profile = user.profile
        writer.writerow(['Role', profile.get_role_display()])
        writer.writerow(['Department', profile.department or 'N/A'])
    except UserProfile.DoesNotExist:
        pass

    writer.writerow([])
    writer.writerow(['Access Logs'])
    writer.writerow(['Timestamp', 'Type', 'IP Address', 'Description'])

    for log in AccessLog.objects.filter(user=user).order_by('-timestamp')[:100]:
        writer.writerow([
            log.timestamp,
            log.get_access_type_display(),
            log.ip_address,
            log.description
        ])

    # Update export request
    try:
        export_req = DataExportRequest.objects.get(user=user, status='processing')
        export_req.status = 'completed'
        export_req.completed_at = timezone.now()
        export_req.save()
    except DataExportRequest.DoesNotExist:
        pass


@login_required
@require_http_methods(["GET"])
def download_user_data(request):
    """Download user data as CSV file"""
    output = StringIO()
    writer = csv.writer(output)

    writer.writerow(['Personal Information'])
    writer.writerow(['Field', 'Value'])
    writer.writerow(['Username', request.user.username])
    writer.writerow(['Email', request.user.email])
    writer.writerow(['First Name', request.user.first_name])
    writer.writerow(['Last Name', request.user.last_name])
    writer.writerow(['Joined Date', request.user.date_joined])

    writer.writerow([])
    writer.writerow(['Access Logs (Last 100)'])
    writer.writerow(['Timestamp', 'Type', 'IP Address', 'Description'])

    for log in AccessLog.objects.filter(user=request.user).order_by('-timestamp')[:100]:
        writer.writerow([
            log.timestamp,
            log.get_access_type_display(),
            log.ip_address,
            log.description
        ])

    # Create response
    response = FileResponse(BytesIO(output.getvalue().encode()))
    response['Content-Type'] = 'text/csv'
    response['Content-Disposition'] = f'attachment; filename="user_data_{request.user.username}.csv"'

    log_access(
        request.user,
        'data_download',
        'Downloaded personal data',
        request
    )

    return response


@login_required
@require_http_methods(["GET", "POST"])
def request_account_deletion(request):
    """Request account deletion"""
    if request.method == 'POST':
        reason = request.POST.get('reason', '')

        # Check if there's already a pending deletion request
        existing_request = DeleteAccountRequest.objects.filter(
            user=request.user,
            status__in=['pending', 'confirmed']
        ).exists()

        if existing_request:
            messages.warning(request, 'You already have a pending account deletion request.')
            return redirect('settings')

        # Create deletion request
        delete_request = DeleteAccountRequest.objects.create(
            user=request.user,
            reason=reason,
            status='pending'
        )

        log_access(
            request.user,
            'data_update',
            'Requested account deletion',
            request
        )

        messages.success(
            request,
            'Account deletion request received. Please check your email to confirm.'
        )

        return redirect('settings')

    return render(request, 'settings.html')


@login_required
@require_http_methods(["POST"])
def confirm_account_deletion(request):
    """Confirm account deletion"""
    try:
        delete_request = DeleteAccountRequest.objects.get(
            user=request.user,
            status='pending'
        )

        delete_request.status = 'confirmed'
        delete_request.confirmed_at = timezone.now()
        delete_request.scheduled_deletion_date = timezone.now() + timedelta(days=30)
        delete_request.save()

        log_access(
            request.user,
            'data_update',
            'Confirmed account deletion (scheduled for 30 days)',
            request
        )

        messages.success(
            request,
            'Account deletion confirmed. Your account will be permanently deleted in 30 days.'
        )

        return redirect('logout')

    except DeleteAccountRequest.DoesNotExist:
        messages.error(request, 'No pending deletion request found.')
        return redirect('settings')


@login_required
@require_http_methods(["POST"])
def cancel_account_deletion(request):
    """Cancel pending account deletion"""
    try:
        delete_request = DeleteAccountRequest.objects.get(
            user=request.user,
            status='confirmed'
        )

        delete_request.status = 'cancelled'
        delete_request.save()

        log_access(
            request.user,
            'data_update',
            'Cancelled account deletion',
            request
        )

        messages.success(request, 'Account deletion has been cancelled.')
        return redirect('settings')

    except DeleteAccountRequest.DoesNotExist:
        messages.error(request, 'No pending deletion request found.')
        return redirect('settings')

@login_required
@role_required('super_admin', 'admin', 'doctor')
def analytics_dashboard(request):
    """Analytics dashboard with KPIs and visualizations"""
    from django.db.models import Count, Q
    from datetime import datetime, timedelta

    # Calculate date ranges
    today = timezone.now().date()

    # Optimize: Use aggregation instead of multiple count() calls
    appointment_stats = PatientAppointment.objects.aggregate(
        total=Count('id'),
        pending=Count('id', filter=Q(status='pending')),
        confirmed=Count('id', filter=Q(status='confirmed')),
        completed=Count('id', filter=Q(status='completed')),
    )

    # KPIs
    total_appointments = appointment_stats['total']
    pending_appointments = appointment_stats['pending']
    confirmed_appointments = appointment_stats['confirmed']
    completed_appointments = appointment_stats['completed']

    # Monthly trend data - optimized with values and annotate
    monthly_data = []
    for i in range(11, -1, -1):
        month_date = today - timedelta(days=30 * i)
        month_start = month_date.replace(day=1)
        if i > 0:
            next_month = (month_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            month_end = next_month - timedelta(days=1)
        else:
            month_end = today

        count = PatientAppointment.objects.filter(
            appointment_date__gte=month_start,
            appointment_date__lte=month_end
        ).count()

        monthly_data.append({
            'month': month_start.strftime('%b'),
            'count': count
        })

    # Appointment types distribution
    type_data = PatientAppointment.objects.values('appointment_type').annotate(
        count=Count('id')
    )

    # Status distribution
    status_data = PatientAppointment.objects.values('status').annotate(
        count=Count('id')
    )

    # Department workload (based on doctor's department)
    department_data = User.objects.filter(
        profile__role='doctor'
    ).values(
        'profile__department'
    ).annotate(
        count=Count('assigned_appointments')
    ).order_by('-count')[:5]

    context = {
        'total_appointments': total_appointments,
        'pending_appointments': pending_appointments,
        'confirmed_appointments': confirmed_appointments,
        'completed_appointments': completed_appointments,
        'monthly_data': json.dumps(list(monthly_data)),
        'type_data': json.dumps(list(type_data)),
        'status_data': json.dumps(list(status_data)),
        'department_data': json.dumps(list(department_data)),
    }

    log_access(request.user, 'data_view', 'Accessed analytics dashboard', request)

    return render(request, 'analytics_dashboard.html', context)

@login_required
def generate_report(request):
    """Generate and download reports"""
    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        format_type = request.POST.get('format', 'csv')
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')

        # Create report record
        from .models import Report  # Import the Report model

        report = Report.objects.create(
            user=request.user,
            report_type=report_type,
            format=format_type,
            date_from=date_from if date_from else None,
            date_to=date_to if date_to else None,
            status='processing',
            expires_at=timezone.now() + timedelta(days=7)
        )

        # Generate report based on type
        try:
            if report_type == 'appointments':
                report_data = generate_appointments_report(date_from, date_to)
            elif report_type == 'patient_records':
                report_data = generate_patient_records_report(date_from, date_to)
            elif report_type == 'analytics':
                report_data = generate_analytics_report(date_from, date_to)
            elif report_type == 'audit':
                report_data = generate_audit_report(date_from, date_to)
            else:
                raise ValueError('Invalid report type')

            # Export based on format
            if format_type == 'csv':
                file_content = export_to_csv(report_data)
                content_type = 'text/csv'
                filename = f'{report_type}_{timezone.now().strftime("%Y%m%d")}.csv'
            elif format_type == 'excel':
                file_content = export_to_excel(report_data)
                content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                filename = f'{report_type}_{timezone.now().strftime("%Y%m%d")}.xlsx'
            elif format_type == 'pdf':
                file_content = export_to_pdf(report_data)
                content_type = 'application/pdf'
                filename = f'{report_type}_{timezone.now().strftime("%Y%m%d")}.pdf'

            report.status = 'completed'
            report.completed_at = timezone.now()
            report.save()

            log_access(request.user, 'data_download', f'Generated {report_type} report', request)

            # Return file
            response = HttpResponse(file_content, content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            report.status = 'failed'
            report.save()
            messages.error(request, f'Error generating report: {str(e)}')
            return redirect('reports')

    # GET request - show report generation form
    context = {
        'report_types': [
            ('appointments', 'Appointments Report'),
            ('patient_records', 'Patient Records Report'),
            ('analytics', 'System Analytics Report'),
            ('audit', 'Audit Trail Report'),
        ],
        'formats': [
            ('csv', 'CSV'),
            ('excel', 'Excel (XLSX)'),
            ('pdf', 'PDF'),
        ]
    }
    return render(request, 'analytics_dashboard.html', context)


def generate_appointments_report(date_from, date_to):
    """Generate appointments report data"""
    appointments = PatientAppointment.objects.all()

    if date_from:
        appointments = appointments.filter(appointment_date__gte=date_from)
    if date_to:
        appointments = appointments.filter(appointment_date__lte=date_to)

    data = []
    for appt in appointments:
        data.append({
            'Date': appt.appointment_date.strftime('%Y-%m-%d'),
            'Time': appt.appointment_time.strftime('%H:%M') if appt.appointment_time else 'N/A',
            'Patient': f'{appt.first_name} {appt.last_name}',
            'Email': appt.email,
            'Contact': appt.contact_number,
            'Type': appt.get_appointment_type_display(),
            'Status': appt.get_status_display(),
            'Doctor': appt.assigned_doctor.get_full_name() if appt.assigned_doctor else 'Unassigned',
            'Created': appt.created_at.strftime('%Y-%m-%d %H:%M'),
        })

    return data


def generate_patient_records_report(date_from, date_to):
    """Generate patient records report data"""
    appointments = PatientAppointment.objects.all()

    if date_from:
        appointments = appointments.filter(created_at__date__gte=date_from)
    if date_to:
        appointments = appointments.filter(created_at__date__lte=date_to)

    # Group by patient
    patient_data = {}
    for appt in appointments:
        patient_key = f"{appt.first_name} {appt.last_name}"
        if patient_key not in patient_data:
            patient_data[patient_key] = {
                'Name': patient_key,
                'Email': appt.email,
                'Contact': appt.contact_number,
                'DOB': appt.date_of_birth.strftime('%Y-%m-%d'),
                'Gender': appt.get_gender_display(),
                'Address': appt.address,
                'Total_Appointments': 0,
                'Last_Visit': None
            }

        patient_data[patient_key]['Total_Appointments'] += 1
        if patient_data[patient_key]['Last_Visit'] is None or appt.appointment_date > patient_data[patient_key][
            'Last_Visit']:
            patient_data[patient_key]['Last_Visit'] = appt.appointment_date.strftime('%Y-%m-%d')

    return list(patient_data.values())


def generate_analytics_report(date_from, date_to):
    """Generate system analytics report data"""
    from django.db.models import Count, Q

    appointments = PatientAppointment.objects.all()

    if date_from:
        appointments = appointments.filter(appointment_date__gte=date_from)
    if date_to:
        appointments = appointments.filter(appointment_date__lte=date_to)

    # Calculate statistics
    stats = appointments.aggregate(
        total=Count('id'),
        pending=Count('id', filter=Q(status='pending')),
        assigned=Count('id', filter=Q(status='assigned')),
        confirmed=Count('id', filter=Q(status='confirmed')),
        completed=Count('id', filter=Q(status='completed')),
        cancelled=Count('id', filter=Q(status='cancelled')),
    )

    # Appointment types
    type_stats = appointments.values('appointment_type').annotate(count=Count('id'))

    data = [
        {'Metric': 'Total Appointments', 'Value': stats['total']},
        {'Metric': 'Pending Appointments', 'Value': stats['pending']},
        {'Metric': 'Assigned Appointments', 'Value': stats['assigned']},
        {'Metric': 'Confirmed Appointments', 'Value': stats['confirmed']},
        {'Metric': 'Completed Appointments', 'Value': stats['completed']},
        {'Metric': 'Cancelled Appointments', 'Value': stats['cancelled']},
        {'Metric': '', 'Value': ''},  # Separator
        {'Metric': 'Appointment Types:', 'Value': ''},
    ]

    for type_stat in type_stats:
        data.append({
            'Metric': f"  {dict(PatientAppointment.APPOINTMENT_TYPE_CHOICES).get(type_stat['appointment_type'], type_stat['appointment_type'])}",
            'Value': type_stat['count']
        })

    return data


def generate_audit_report(date_from, date_to):
    """Generate audit trail report data"""
    logs = AccessLog.objects.all()

    if date_from:
        logs = logs.filter(timestamp__date__gte=date_from)
    if date_to:
        logs = logs.filter(timestamp__date__lte=date_to)

    logs = logs.order_by('-timestamp')[:500]  # Limit to 500 most recent

    data = []
    for log in logs:
        data.append({
            'Timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'User': log.user.username,
            'User_Email': log.user.email,
            'Access_Type': log.get_access_type_display(),
            'IP_Address': log.ip_address or 'N/A',
            'Description': log.description or 'N/A',
        })

    return data


def export_to_csv(data):
    """Export data to CSV format"""
    output = StringIO()

    if data:
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

    return output.getvalue().encode('utf-8')


def export_to_excel(data):
    """Export data to Excel format"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    if data:
        # Headers
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data.values(), 1):
                ws.cell(row=row_num, column=col_num, value=str(value))

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def export_to_pdf(data):
    """Export data to PDF format"""
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    title = Paragraph("<b>System Report</b>", styles['Title'])
    elements.append(title)

    if data:
        # Convert data to table format
        table_data = [list(data[0].keys())]  # Headers
        for row in data:
            table_data.append([str(v) for v in row.values()])

        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(table)

    doc.build(elements)
    return output.getvalue()